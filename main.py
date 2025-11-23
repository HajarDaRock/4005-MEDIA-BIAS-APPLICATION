from fastapi import FastAPI, Request
from fastapi.responses import JSONResponse, HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

from article_utils import fetch_article, is_restricted_url
from classify_articles import classify_bias

import pandas as pd
from openpyxl import load_workbook
import os
import asyncio

app = FastAPI()

app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

# Constants for Excel file handling
EXCEL_PATH = "ExtractedData.xlsx"
EXCEL_SHEET = "BiasResults"

@app.get("/", response_class=HTMLResponse)
async def serve_index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})

@app.get("/a", response_class=HTMLResponse)
async def serve_about(request: Request):
    return templates.TemplateResponse("AboutUs.html", {"request": request})

# Enable CORS for frontend integration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow all origins
    allow_methods=["*"],  # Allow all HTTP methods
    allow_headers=["*"],  # Allow all headers
)

@app.post("/classify")
async def classify_url(request: Request):
    try:
        data = await request.json()
        url = data.get("url")
        print(f"[INPUT] URL received: {url}")

        if not url:
            return JSONResponse(content={"error": "No URL provided."}, status_code=400)
        if not url.startswith("http"):
            return JSONResponse(
                content={"error": "Invalid URL format. Must start with http or https."},
                status_code=400
            )

        if is_restricted_url(url):
            print(f"[BLOCKED] Restricted domain: {url}")
            return JSONResponse(
                content={"error": "Access denied: this outlet is not supported."},
                status_code=403
            )

        title, content = fetch_article(url)
        print(f"[SCRAPE] Title: {title}")
        print(f"[SCRAPE] Content length: {len(content) if content else 'None'}")

        if not content:
            return JSONResponse(
                content={"error": "Unable to extract content from the article."},
                status_code=400
            )

        loop = asyncio.get_event_loop()
        bias = await loop.run_in_executor(None, classify_bias, content)
        print(f"[MODEL] Raw bias result: {repr(bias)}")

        # Default to 'Neutral' if bias is empty
        if not bias:
            bias = "Neutral"
            print("[MODEL] Bias was empty or null; defaulting to 'Neutral'.")

        new_data = pd.DataFrame([{
            "URL": url,
            "Title": title,
            "Content": content,
            "Bias": bias
        }])

        # Appends data to an existing Excel file or creates a new one.
        # This is for basic data logging of predictions.
        try:
            if os.path.exists(EXCEL_PATH):
                book = load_workbook(EXCEL_PATH)
                with pd.ExcelWriter(EXCEL_PATH, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    if EXCEL_SHEET in book.sheetnames:
                        startrow = book[EXCEL_SHEET].max_row
                        new_data.to_excel(writer, sheet_name=EXCEL_SHEET, index=False, header=False, startrow=startrow)
                    else:
                        new_data.to_excel(writer, sheet_name=EXCEL_SHEET, index=False)
                print("[EXCEL] Data appended to existing file.")
            else:
                with pd.ExcelWriter(EXCEL_PATH, engine='openpyxl') as writer:
                    new_data.to_excel(writer, sheet_name=EXCEL_SHEET, index=False)
                print("[EXCEL] New Excel file created and data saved.")
        except Exception as e:
            print(f"[EXCEL ERROR] {e}")
            return JSONResponse(
                content={"error": f"Bias classified, but failed to save to Excel: {str(e)}"},
                status_code=500
            )

        return {"bias": bias, "title": title}

    # Catch-all for unexpected errors
    except Exception as e:
        print(f"[SERVER ERROR] {e}")
        return JSONResponse(
            content={"error": f"Unexpected error: {str(e)}"},
            status_code=500
        )
