"""
A command-line utility for classifying the political bias of news articles.

This script allows a user to classify articles from URLs provided via command-line
arguments, a text file, or interactive input. The results, including the article's
title, content, and predicted bias, are appended to specified sheets in an
Excel workbook.

It serves as a tool for manual testing and bulk processing without needing to run
the FastAPI frontend.
"""

import argparse
import os
from typing import List

import pandas as pd
from openpyxl import load_workbook

from article_utils import fetch_article, is_restricted_url
from classify_articles import classify_bias


def parse_args() -> argparse.Namespace:
    """Defines and parses command-line arguments for the script."""
    parser = argparse.ArgumentParser(description="Classify article bias without running the API.")
    parser.add_argument(
        "--url",
        action="append",
        help="Article URL to process. Use multiple --url flags for multiple links.",
    )
    parser.add_argument(
        "--file",
        help="Path to a text file containing one article URL per line.",
    )
    parser.add_argument(
        "--excel",
        default="ExtractedData.xlsx",
        help="Path to the Excel workbook to append results to (default: ExtractedData.xlsx).",
    )
    parser.add_argument(
        "--full-sheet",
        default="Sheet1",
        help="Sheet name for detailed rows (default: Sheet1).",
    )
    parser.add_argument(
        "--summary-sheet",
        default="Sheet2",
        help="Sheet name for summary rows (default: Sheet2).",
    )
    return parser.parse_args()


def load_urls_from_file(path: str) -> List[str]:
    """Loads a list of URLs from a text file, one URL per line."""
    urls: List[str] = []
    if not os.path.exists(path):
        raise FileNotFoundError(f"URL file not found: {path}")
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            value = line.strip()
            if value:
                urls.append(value)
    return urls


def resolve_urls(args: argparse.Namespace) -> List[str]:
    """
    Consolidates URLs from command-line arguments and/or a file.
    Falls back to interactive input if no URLs are provided.
    """
    urls: List[str] = []
    if args.url:
        urls.extend(u.strip() for u in args.url if u.strip())
    if args.file:
        urls.extend(load_urls_from_file(args.file))

    if urls:
        return urls

    print("No URLs provided via --url or --file. Enter URLs manually (blank line to finish):")
    while True:
        value = input("> ").strip()
        if not value:
            break
        urls.append(value)
    return urls


def write_to_excel(full_df: pd.DataFrame, summary_df: pd.DataFrame, excel_path: str, full_sheet: str, summary_sheet: str) -> None:
    """
    Appends DataFrames to specified sheets in an Excel file.

    If the file or sheets do not exist, they are created. If they do exist,
    data is appended to the next available row.
    """
    if full_df.empty:
        print("No rows to persist; skipping Excel update.")
        return

    if os.path.exists(excel_path):
        book = load_workbook(excel_path)
        with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            if full_sheet in book.sheetnames:
                startrow_full = book[full_sheet].max_row
                full_df.to_excel(writer, sheet_name=full_sheet, index=False, header=startrow_full == 1, startrow=startrow_full)
            else:
                full_df.to_excel(writer, sheet_name=full_sheet, index=False)

            if summary_sheet in book.sheetnames:
                startrow_summary = book[summary_sheet].max_row
                summary_df.to_excel(writer, sheet_name=summary_sheet, index=False, header=startrow_summary == 1, startrow=startrow_summary)
            else:
                summary_df.to_excel(writer, sheet_name=summary_sheet, index=False)
    else:
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            full_df.to_excel(writer, sheet_name=full_sheet, index=False)
            summary_df.to_excel(writer, sheet_name=summary_sheet, index=False)


def main() -> int:
    """
    Main function to orchestrate the article classification process.
    
    - Parses arguments.
    - Resolves URLs to process.
    - Fetches, classifies, and collects data for each URL.
    - Writes the collected data to an Excel file.
    """
    args = parse_args()
    urls = resolve_urls(args)
    if not urls:
        print("No URLs supplied. Exiting.")
        return 0

    new_full_data = []
    new_summary_data = []

    for url in urls:
        if is_restricted_url(url):
            print(f"Access restricted: {url}")
            continue

        title, content = fetch_article(url)
        if not (title and content):
            print(f"Unable to fetch article content: {url}")
            continue

        raw_response = classify_bias(content) or ""
        bias = raw_response.strip().capitalize() if raw_response.strip() else "Unknown"

        print(f"\nURL: {url}")
        print(f"Raw model output: '{raw_response}'")
        print(f"Final bias classification: {bias}\n")

        new_full_data.append(
            {"Title": title, "Content": content, "URL": url, "Bias": bias, "RawOutput": raw_response}
        )
        new_summary_data.append({"Title": title, "Bias": bias, "RawOutput": raw_response})

    df_full = pd.DataFrame(new_full_data)
    df_summary = pd.DataFrame(new_summary_data)
    write_to_excel(df_full, df_summary, args.excel, args.full_sheet, args.summary_sheet)
    print(f"\nProcessed {len(new_full_data)} articles. Results saved to '{args.excel}'.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
